import openpyxl


def get_rowcount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row


def get_colcount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column


def getcelldata(path, sheetname, rowno, colno):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowno, column=colno).value


def setcelldata(path, sheetname, rowno, colno, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row=rowno, column=colno).value = data
    workbook.save(path)


path="C:\\Users\\VARUN\\Desktop\\Varun_Personal\\Selenium_Python_Arora\\data_1.xlsx"
sheetname="Sheet1"

#calling methods
rows=get_rowcount(path,sheetname)
print(rows) # 7

cols=get_colcount(path,sheetname)
print(cols) # 3

print(getcelldata(path,sheetname,2,1)) # trainer@automation.com

setcelldata(path,sheetname,1,4,'DOB')
